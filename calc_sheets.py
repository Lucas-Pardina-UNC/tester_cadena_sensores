import os
import openpyxl  #pip install openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.exceptions import InvalidFileException
from typing import List, Tuple, Optional

def center_and_adjust_columns(sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """
    Centers the content of all cells that have data and adjusts the column widths
    to fit the content in the specified sheet.
    
    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The Excel worksheet to modify.
    
    Returns:
        None
    """
    # Center-align content of all cells that have data
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths to fit the content
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the letter of the column
        
        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        
        # Set the width with some padding
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width

def save_workbook_safely(file_path: str, workbook: openpyxl.workbook.workbook.Workbook) -> bool:
    """
    Attempts to save the workbook to the specified file path, handling errors like 
    permission issues or invalid file formats.
    
    Parameters:
        file_path (str): The path to save the workbook to.
        workbook (openpyxl.workbook.workbook.Workbook): The workbook to save.
    
    Returns:
        bool: True if the file was saved successfully, False otherwise.
    """
    while True:
        try:
            workbook.save(file_path)
            #print(f"Archivo guardado exitosamente en '{file_path}'.")
            return True  # Return True on success
        except PermissionError:
            print(f"Permiso denegado: Cierra el archivo de Excel '{file_path}' e intenta nuevamente.")
            input("Presiona Enter cuando el archivo esté cerrado...")  # Wait for user confirmation
        except InvalidFileException as e:
            print(f"Error al guardar el archivo: {e}")
            return False  # Return False on error

def log_to_excel(log_data: List[Tuple], timestamp: Optional[str] = None, chain_id: int = 0, filename: str = "modbus_test_log.xlsx") -> None:
    """
    Logs data to an Excel file with columns for timestamp, Slave ID, ADC value, and temperature,
    starting in a specified column set based on chain_id.
    
    Parameters:
        log_data (List[Tuple]): A list of tuples where each tuple contains the data to log (Slave ID, ADC value, Temperature).
        timestamp (Optional[str]): The timestamp to log. Defaults to None.
        chain_id (int): An integer that determines where the data will be logged. Defaults to 0.
        filename (str): The name of the Excel file to log the data to. Defaults to "modbus_test_log.xlsx".
    
    Returns:
        None
    """
    # Check if the file exists
    if os.path.exists(filename):
        # If the file exists, open it
        workbook = openpyxl.load_workbook(filename)
        #print(f"Archivo existente abierto: {filename}")
    else:
        # If the file does not exist, create a new workbook and add a sheet
        workbook = openpyxl.Workbook()
        print(f"Archivo nuevo creado: {filename}")
        # Select the active sheet
        sheet = workbook.active
        # Rename the active sheet
        sheet.title = "Log Data"
        
    # Select the active sheet
    sheet = workbook.active

    # Calculate the starting column based on chain_id (0 -> A, 1 -> F, 2 -> K, etc.), with each dataset taking 5 columns
    start_column = 1 + (chain_id * 5)  # Each chain_id starts 5 columns apart
    
    # Write headers if they're not present in the specified starting column
    if sheet.cell(row=1, column=start_column).value != "Timestamp":
        sheet.cell(row=1, column=start_column, value="Timestamp")  
        sheet.cell(row=1, column=start_column + 1, value="Slave_ID")  
        sheet.cell(row=1, column=start_column + 2, value="ADC_Value")  
        sheet.cell(row=1, column=start_column + 3, value="Temperature") 

    # Find the first empty row within the specific columns for the chain_id
    data_write_row = 2  # Start after header row
    while sheet.cell(row=data_write_row, column=start_column).value is not None or \
          sheet.cell(row=data_write_row, column=start_column + 1).value is not None or \
          sheet.cell(row=data_write_row, column=start_column + 2).value is not None or \
          sheet.cell(row=data_write_row, column=start_column + 3).value is not None:
        data_write_row += 1

    # Write each entry from log_data to the specified row and columns
    for entry in log_data:
        sheet.cell(row=data_write_row, column=start_column, value=timestamp)  # Timestamp
        sheet.cell(row=data_write_row, column=start_column + 1, value=entry[0])  # Slave_ID
        sheet.cell(row=data_write_row, column=start_column + 2, value=entry[1])  # ADC Value
        sheet.cell(row=data_write_row, column=start_column + 3, value=entry[2])  # Temperature
        data_write_row += 1  # Move to the next row for each log entry

    # Format cells
    center_and_adjust_columns(sheet)

    # Save the workbook
    save_workbook_safely(filename, workbook)
    #print(f"Datos registrados en {filename} en las columnas comenzando en {chr(64 + start_column)}")
